#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync.py — Régénère les fichiers data/*.js depuis admin.xlsx
Usage   : python sync.py
Résultat: écrase les 5 fichiers data/*.js avec le contenu d'admin.xlsx.

IMPORTANT : fermer Excel avant de lancer ce script.
"""

import json, os, sys
from pathlib import Path
from datetime import datetime

# ── Auto-install openpyxl si absent ──────────────────────────────
try:
    import openpyxl
except ImportError:
    print("  openpyxl non trouvé — installation en cours...")
    os.system(f'"{sys.executable}" -m pip install openpyxl')
    import openpyxl

ROOT  = Path(__file__).parent
DATA  = ROOT / "data"
EXCEL = ROOT / "admin.xlsx"


# ─────────────────────────────────────────────────────────────────
# Helpers lecture
# ─────────────────────────────────────────────────────────────────

def sv(val, default=""):
    """Valeur safe : None / vide → default."""
    if val is None:
        return default
    v = str(val).strip()
    return v if v else default


def bv(val):
    """OUI/NON → True/False."""
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    v = str(val).strip().upper()
    if v in ("OUI", "TRUE", "1", "YES", "VRAI"):
        return True
    return False


def iv(val, default=0):
    """Entier safe."""
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def read_sheet(wb, name):
    """Retourne les lignes non-vides d'une feuille (sans l'en-tête)."""
    if name not in wb.sheetnames:
        print(f"  AVERTISSEMENT : feuille '{name}' introuvable dans admin.xlsx")
        return []
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        if any(v is not None and str(v).strip() for v in r):
            rows.append(r)
    return rows


def pad(row, n):
    """Complète une ligne jusqu'à n colonnes."""
    r = list(row)
    while len(r) < n:
        r.append(None)
    return r


# ─────────────────────────────────────────────────────────────────
# Écriture des fichiers JS
# ─────────────────────────────────────────────────────────────────

HEADER_SITE = """\
// =====================================================
// FICHIER GLOBAL DU SITE
// Généré automatiquement par sync.py le {date}
// Pour modifier via Excel : admin.xlsx > onglet Config / Navigation
// Ou modifier directement ce fichier (les deux sont compatibles).
// Ne pas renommer window.BDE_SITE.
// =====================================================
"""

HEADER_EVENTS = """\
// =====================================================
// ÉVÉNEMENTS
// Généré automatiquement par sync.py le {date}
// Pour modifier via Excel : admin.xlsx > onglets Events / Event_*
// Ou modifier directement ce fichier (les deux sont compatibles).
// Ne pas renommer window.BDE_EVENTS.
// =====================================================
"""

HEADER_SPONSORS = """\
// =====================================================
// SPONSORS / PARTENAIRES
// Généré automatiquement par sync.py le {date}
// Pour modifier via Excel : admin.xlsx > onglet Sponsors
// Ou modifier directement ce fichier (les deux sont compatibles).
// Ne pas renommer window.BDE_SPONSORS.
// =====================================================
"""

HEADER_ARTISTS = """\
// =====================================================
// ARTISTES / SHOWCASES
// Généré automatiquement par sync.py le {date}
// Pour modifier via Excel : admin.xlsx > onglets Artists_Cartes / Artists_Bande
// Ou modifier directement ce fichier (les deux sont compatibles).
// Ne pas renommer window.BDE_ARTISTS.
// =====================================================
"""

HEADER_GALLERIES = """\
// =====================================================
// GALERIES ET ALBUMS GOOGLE PHOTOS
// Généré automatiquement par sync.py le {date}
// Pour modifier via Excel : admin.xlsx > onglets Galeries / Galerie_Images
// Ou modifier directement ce fichier (les deux sont compatibles).
// Ne pas renommer window.BDE_GALLERIES.
// =====================================================
"""


def write_js(filename, var_name, data, header_tpl):
    path = DATA / filename
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    header = header_tpl.format(date=date_str)
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    content = f"{header}\nwindow.{var_name} = {json_str};\n"
    path.write_text(content, encoding="utf-8")
    n = len(data) if isinstance(data, list) else (
        len(data.get("cards", [])) if isinstance(data, dict) else "—"
    )
    print(f"    OK  {filename}  ({n} elements)")


# ─────────────────────────────────────────────────────────────────
# Builders
# ─────────────────────────────────────────────────────────────────

def build_site(wb):
    cfg_rows = read_sheet(wb, "Config")
    nav_rows = read_sheet(wb, "Navigation")

    # Config → dict clé/valeur
    cfg = {}
    for row in cfg_rows:
        row = pad(row, 2)
        key = sv(row[0])
        val = sv(row[1])
        if key:
            cfg[key] = val

    def g(k, default=""):
        return cfg.get(k, default)

    # Navigation
    nav = []
    for row in sorted(nav_rows, key=lambda r: iv(r[0])):
        row = pad(row, 5)
        label    = sv(row[1])
        label_en = sv(row[2])
        href     = sv(row[3])
        ext      = sv(row[4])
        if not label and not href:
            continue
        item = {"label": label, "labelEn": label_en, "href": href}
        if ext:
            item["externalFromNonHome"] = ext
        nav.append(item)

    site = {
        "siteName":             g("siteName"),
        "university":           g("university"),
        "year":                 g("year"),
        "seasonLabel":          g("seasonLabel"),
        "email":                g("email"),
        "phone":                g("phone"),
        "phoneHref":            g("phoneHref"),
        "address":              g("address"),
        "logo":                 g("logo"),
        "footerDescription":    g("footerDescription"),
        "footerDescriptionEn":  g("footerDescriptionEn"),
        "instagramUrl":         g("instagramUrl"),
        "instagramLabel":       g("instagramLabel"),
        "nuitsInstagramUrl":    g("nuitsInstagramUrl"),
        "nuitsInstagramLabel":  g("nuitsInstagramLabel"),
        "tiktokUrl":            g("tiktokUrl"),
        "facebookUrl":          g("facebookUrl"),
        "defaultTicketUrl":     g("defaultTicketUrl"),
        "defaultGooglePhotosUrl": g("defaultGooglePhotosUrl"),
        "copyrightText":        g("copyrightText"),
        "ticket": {
            "eventName":      g("ticket_eventName"),
            "eventSub":       g("ticket_eventSub"),
            "eventSubEn":     g("ticket_eventSubEn"),
            "url":            g("ticket_url"),
            "buttonLabel":    g("ticket_buttonLabel"),
            "buttonLabelEn":  g("ticket_buttonLabelEn"),
            "description":    g("ticket_description"),
            "descriptionEn":  g("ticket_descriptionEn"),
            "note":           g("ticket_note"),
            "noteEn":         g("ticket_noteEn"),
        },
        "navigation": nav,
    }

    write_js("site.js", "BDE_SITE", site, HEADER_SITE)


def build_events(wb):
    ev_rows   = read_sheet(wb, "Events")
    desc_rows = read_sheet(wb, "Event_Descriptions")
    meta_rows = read_sheet(wb, "Event_Meta")
    tag_rows  = read_sheet(wb, "Event_Tags")
    art_rows  = read_sheet(wb, "Event_Artists")

    # Indexer les sous-tables par slug
    def index_by_slug(rows, slug_col=0, order_col=1):
        idx = {}
        for r in rows:
            r = pad(r, max(slug_col, order_col) + 2)
            slug = sv(r[slug_col])
            if slug:
                idx.setdefault(slug, []).append(r)
        for k in idx:
            idx[k].sort(key=lambda r: iv(r[order_col]))
        return idx

    descs   = index_by_slug(desc_rows)
    metas   = index_by_slug(meta_rows)
    tags_ix = index_by_slug(tag_rows)
    arts    = index_by_slug(art_rows)

    events = []
    for row in ev_rows:
        row = pad(row, 37)
        slug = sv(row[0])
        if not slug:
            continue

        # Image principale + images supplémentaires
        main_img = sv(row[19])
        supp_raw = sv(row[20])
        supp = [x.strip() for x in supp_raw.split(";") if x.strip()] if supp_raw else []
        # Reconstituer la liste complète sans doublon
        images = [main_img] if main_img else []
        for img in supp:
            if img not in images:
                images.append(img)

        # Blocs de description
        desc_blocks = []
        for dr in descs.get(slug, []):
            dr = pad(dr, 4)
            fr_text = sv(dr[2])
            en_text = sv(dr[3])
            if fr_text or en_text:
                desc_blocks.append({
                    "text": fr_text,
                    "i18n": {"fr": fr_text, "en": en_text},
                })

        # Meta
        meta_out = []
        for mr in metas.get(slug, []):
            mr = pad(mr, 6)
            cle_fr = sv(mr[2])
            cle_en = sv(mr[3])
            val_fr = sv(mr[4])
            val_en = sv(mr[5])
            if not cle_fr:
                continue
            meta_out.append({
                "key":      cle_fr,
                "value":    val_fr,
                "keyI18n":  {"fr": cle_fr, "en": cle_en},
                "valueI18n": {"fr": val_fr, "en": val_en} if val_en else {},
            })

        # Tags
        tags_out = []
        for tr in tags_ix.get(slug, []):
            tr = pad(tr, 6)
            label  = sv(tr[2])
            rouge  = bv(tr[3])
            lbl_fr = sv(tr[4]) or label
            lbl_en = sv(tr[5])
            if not label:
                continue
            tag_item = {"label": label, "red": rouge}
            if lbl_fr or lbl_en:
                tag_item["i18n"] = {"fr": lbl_fr, "en": lbl_en}
            tags_out.append(tag_item)

        # Artistes liés à l'événement
        artists_out = []
        for ar in arts.get(slug, []):
            ar = pad(ar, 4)
            lbl = sv(ar[2])
            hl  = bv(ar[3])
            if lbl:
                artists_out.append({"label": lbl, "highlight": hl})

        hp_fr = sv(row[17])
        hp_en = sv(row[18])

        ev = {
            "slug":     slug,
            "order":    iv(row[1], 1),
            "number":   sv(row[2]),
            "title":    sv(row[3]),
            "subtitle": sv(row[4]),
            "category": sv(row[5]),
            "date":     sv(row[6]),
            "dateLabel":   {"fr": sv(row[7]),  "en": sv(row[8])},
            "statusLabel": {"fr": sv(row[9]),  "en": sv(row[10])},
            "statusColor": sv(row[11]),
            "place":    sv(row[12]),
            "venue":    sv(row[13]),
            "shortDescription":   sv(row[14]),
            "homeDescriptionI18n": {"fr": sv(row[15]), "en": sv(row[16])},
            "homePeriod":     hp_fr,
            "homePeriodI18n": {"fr": hp_fr, "en": hp_en},
            # longDescription : dérivé des blocs pour compatibilité
            "longDescription": [b["text"] for b in desc_blocks],
            "descriptionBlocks": desc_blocks,
            "image":  main_img,
            "images": images,
            "alt":    sv(row[21]),
            "ticketUrl":   sv(row[22]),
            "ticketLabel": sv(row[23]),
            "dossierUrl":  sv(row[24]),
            "dossierLabel": {"fr": sv(row[25]), "en": sv(row[26])},
            "dossierDownload": bv(row[27]),
            "googlePhotosUrl": sv(row[28]),
            "galleryPage":     sv(row[29]),
            "galleryLabel":    sv(row[30]),
            "showOnHome":       bv(row[31]),
            "showOnEventsPage": bv(row[32]),
            "showOnNuitsPage":  bv(row[33]),
            "featured": bv(row[34]),
            "status":   sv(row[35]),
            "tags":     tags_out,
            "meta":     meta_out,
            "artists":  artists_out,
            "reverse":  bv(row[36]),
        }
        events.append(ev)

    events.sort(key=lambda e: e["order"])
    write_js("events.js", "BDE_EVENTS", events, HEADER_EVENTS)


def build_sponsors(wb):
    rows = read_sheet(wb, "Sponsors")
    sponsors = []
    for row in rows:
        row = pad(row, 11)
        name = sv(row[0])
        if not name:
            continue
        sponsors.append({
            "name":      name,
            "logo":      sv(row[1]),
            "fallback":  sv(row[2]),
            "alt":       sv(row[3]),
            "url":       sv(row[4]),
            "category":  sv(row[5]),
            "type":      sv(row[6]),
            "typeEn":    sv(row[7]),
            "linkLabel": sv(row[8]),
            "active":    bv(row[9]),
            "order":     iv(row[10], 1),
        })
    sponsors.sort(key=lambda x: x["order"])
    write_js("sponsors.js", "BDE_SPONSORS", sponsors, HEADER_SPONSORS)


def build_artists(wb):
    card_rows = read_sheet(wb, "Artists_Cartes")
    band_rows = read_sheet(wb, "Artists_Bande")

    cards = []
    for row in card_rows:
        row = pad(row, 9)
        name = sv(row[0])
        if not name:
            continue
        cards.append({
            "name":       name,
            "image":      sv(row[1]),
            "alt":        sv(row[2]),
            "yearEvent":  sv(row[3]),
            "eventText":  sv(row[4]),
            "badge":      sv(row[5]),
            "featured":   bv(row[6]),
            "active":     bv(row[7]),
            "order":      iv(row[8], 1),
        })
    cards.sort(key=lambda c: c["order"])

    text_strip = []
    for row in sorted(band_rows, key=lambda r: iv(pad(r, 2)[0])):
        row = pad(row, 2)
        name = sv(row[1]) or sv(row[0])
        if name:
            text_strip.append(name)

    write_js("artists.js", "BDE_ARTISTS", {"cards": cards, "textStrip": text_strip}, HEADER_ARTISTS)


def build_galleries(wb):
    gal_rows = read_sheet(wb, "Galeries")
    img_rows = read_sheet(wb, "Galerie_Images")

    # Indexer les images par slug de galerie
    imgs_by_slug = {}
    for row in img_rows:
        row = pad(row, 8)
        slug = sv(row[0])
        if slug:
            imgs_by_slug.setdefault(slug, []).append(row)
    for k in imgs_by_slug:
        imgs_by_slug[k].sort(key=lambda r: iv(pad(r, 2)[1]))

    galleries = []
    for row in gal_rows:
        row = pad(row, 9)
        slug = sv(row[0])
        if not slug:
            continue

        images = []
        for ir in imgs_by_slug.get(slug, []):
            ir = pad(ir, 8)
            src    = sv(ir[2])
            alt    = sv(ir[3])
            cap_fr = sv(ir[4])
            cap_en = sv(ir[5])
            tag_fr = sv(ir[6])
            tag_en = sv(ir[7])
            if not src:
                continue
            images.append({
                "src":     src,
                "alt":     alt,
                "caption": cap_fr,
                "captionI18n": {"fr": cap_fr, "en": cap_en} if cap_en else {},
                "tag":     tag_fr,
                "tagI18n": {"fr": tag_fr, "en": tag_en} if tag_en else {},
            })

        galleries.append({
            "slug":            slug,
            "title":           sv(row[1]),
            "page":            sv(row[2]),
            "eventSlug":       sv(row[3]),
            "coverImage":      sv(row[4]),
            "googlePhotosUrl": sv(row[5]),
            "lightboxMode":    sv(row[6]),
            "active":          bv(row[7]),
            "order":           iv(row[8], 1),
            "images":          images,
        })

    galleries.sort(key=lambda g: g["order"])
    write_js("galleries.js", "BDE_GALLERIES", galleries, HEADER_GALLERIES)


# ─────────────────────────────────────────────────────────────────
# Point d'entrée
# ─────────────────────────────────────────────────────────────────

def main():
    if not EXCEL.exists():
        print()
        print(f"  ERREUR : admin.xlsx introuvable ({EXCEL})")
        print("  Lance d'abord init_excel.bat pour créer le fichier Excel.")
        print()
        sys.exit(1)

    print()
    print("  Lecture de admin.xlsx...")
    try:
        wb = openpyxl.load_workbook(EXCEL, data_only=True)
    except Exception as e:
        print()
        print(f"  ERREUR lors de l'ouverture d'admin.xlsx : {e}")
        print("  Vérifie que le fichier n'est pas ouvert dans Excel.")
        print()
        sys.exit(1)

    print("  Génération des fichiers data/*.js...")
    build_site(wb)
    build_events(wb)
    build_sponsors(wb)
    build_artists(wb)
    build_galleries(wb)

    print()
    print(f"  Terminé ! 5 fichiers data/*.js mis à jour.")
    print(f"  {datetime.now().strftime('%d/%m/%Y  %H:%M:%S')}")
    print()


if __name__ == "__main__":
    main()
