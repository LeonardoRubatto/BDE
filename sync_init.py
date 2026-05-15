#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
sync_init.py — Initialise admin.xlsx depuis les fichiers data/*.js
Usage   : python sync_init.py
Résultat: crée ou écrase admin.xlsx avec toutes les données actuelles du site.
"""

import json, re, os, sys
from pathlib import Path

# ── Auto-install openpyxl si absent ──────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("  openpyxl non trouvé — installation en cours...")
    os.system(f'"{sys.executable}" -m pip install openpyxl')
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

ROOT  = Path(__file__).parent
DATA  = ROOT / "data"
EXCEL = ROOT / "admin.xlsx"

# ── Palette Excel ─────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")   # bleu nuit
HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
ROW1_FILL = PatternFill("solid", fgColor="F0F4FA")   # bleu pâle
ROW2_FILL = PatternFill("solid", fgColor="FFFFFF")   # blanc
KEY_FONT  = Font(bold=True, size=10)
NOTE_FONT = Font(color="999999", italic=True, size=9)


# ─────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────

def load_js(filename: str):
    """Lit un fichier data/*.js et retourne l'objet Python correspondant."""
    path = DATA / filename
    content = path.read_text(encoding="utf-8")
    # Supprimer les lignes de commentaires
    lines = [l for l in content.split("\n") if not l.strip().startswith("//")]
    content = "\n".join(lines)
    # Extraire la valeur après « window.BDE_X = … ; »
    match = re.search(r'window\.\w+\s*=\s*([\s\S]+?)\s*;?\s*$', content.strip())
    if not match:
        raise ValueError(f"Impossible de parser {filename}")
    return json.loads(match.group(1))


def make_header(ws, columns, freeze_row=2):
    """Écrit la rangée d'en-têtes (row 1) et règle les largeurs de colonnes."""
    for col, (label, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)


def row_fill(i):
    """Alternance de couleurs sur les lignes de données."""
    return ROW1_FILL if i % 2 == 0 else ROW2_FILL


def b(val):
    """bool → OUI / NON."""
    if val is True:
        return "OUI"
    if val is False:
        return "NON"
    return val if val is not None else ""


def s(val):
    """Valeur safe : None → chaîne vide."""
    return val if val is not None else ""


def paint_row(ws, row_idx, n_cols):
    fill = row_fill(row_idx)
    for c in range(1, n_cols + 1):
        ws.cell(row_idx, c).fill = fill


# ─────────────────────────────────────────────────────────────────
# Construction des feuilles
# ─────────────────────────────────────────────────────────────────

def sheet_legende(wb):
    ws = wb.create_sheet("LÉGENDE")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 42

    rows = [
        ("FEUILLE", "CONTENU", "NOTES"),
        ("Config", "Paramètres globaux du site (nom, contact, réseaux sociaux, modal billetterie)", "Modifier uniquement la colonne VALEUR"),
        ("Navigation", "Liens du menu de navigation", "Un lien par ligne — respecter l'ordre"),
        ("Events", "Liste des événements (champs principaux)", "Un événement par ligne"),
        ("Event_Descriptions", "Paragraphes de description longue par événement", "Plusieurs lignes par événement (slug_event = clé de liaison)"),
        ("Event_Meta", "Lignes d'info affichées sous chaque événement", "Plusieurs lignes par événement"),
        ("Event_Tags", "Étiquettes/badges de chaque événement", "Plusieurs lignes par événement"),
        ("Event_Artists", "Artistes liés à un événement (line-up)", "Plusieurs lignes par événement"),
        ("Sponsors", "Partenaires/sponsors du BDE", "Un partenaire par ligne"),
        ("Artists_Cartes", "Cartes artistes du carrousel homepage", "Une carte par ligne"),
        ("Artists_Bande", "Noms dans la bande de texte défilant (homepage)", "Un nom par ligne"),
        ("Galeries", "Galeries photos — paramètres principaux", "Une galerie par ligne"),
        ("Galerie_Images", "Photos dans chaque galerie", "Plusieurs lignes par galerie (slug_galerie = clé de liaison)"),
        ("", "", ""),
        ("CONVENTION", "VALEUR À SAISIR", ""),
        ("Booléen VRAI", "OUI", ""),
        ("Booléen FAUX", "NON", ""),
        ("Champ vide", "(laisser la cellule vide)", ""),
        ("Images multiples (colonne images suppl.)", "uploads/a.jpg ; uploads/b.jpg ; uploads/c.jpg", "Séparer chaque chemin par un point-virgule"),
        ("", "", ""),
        ("WORKFLOW EXCEL → SITE", "", ""),
        ("1.", "Modifier les données dans les feuilles ci-dessus", ""),
        ("2.", "Fermer Excel (important)", ""),
        ("3.", "Double-cliquer sur update_site.bat", ""),
        ("4.", "Les fichiers data/*.js sont régénérés automatiquement", ""),
        ("5.", "Déployer sur Cloudflare comme d'habitude", ""),
        ("", "", ""),
        ("WORKFLOW SITE → EXCEL (sync inverse)", "", ""),
        ("1.", "Si tu as modifié directement un fichier data/*.js", ""),
        ("2.", "Fermer Excel si ouvert", ""),
        ("3.", "Double-cliquer sur init_excel.bat pour resynchroniser Excel", ""),
    ]

    for i, (a, b_, c) in enumerate(rows, 1):
        ca = ws.cell(i, 1, a)
        cb = ws.cell(i, 2, b_)
        cc = ws.cell(i, 3, c)
        if i == 1 or a in ("CONVENTION", "WORKFLOW EXCEL → SITE", "WORKFLOW SITE → EXCEL (sync inverse)"):
            for cell in (ca, cb, cc):
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
        elif a and a not in ("", "1.", "2.", "3.", "4.", "5."):
            for cell in (ca, cb, cc):
                cell.fill = row_fill(i)
            ca.font = KEY_FONT


def sheet_config(wb, site):
    ws = wb.create_sheet("Config")
    make_header(ws, [
        ("CHAMP (ne pas modifier)", 32),
        ("VALEUR", 62),
        ("DESCRIPTION", 52),
    ])

    ticket = site.get("ticket") or {}

    rows = [
        ("siteName",              s(site.get("siteName")),               "Nom du site affiché dans les métadonnées"),
        ("university",            s(site.get("university")),             "Nom de l'université"),
        ("year",                  s(site.get("year")),                   "Année en cours (ex: 2026)"),
        ("seasonLabel",           s(site.get("seasonLabel")),            "Libellé de saison (ex: Saison 2025-2026)"),
        ("email",                 s(site.get("email")),                  "Email de contact affiché"),
        ("phone",                 s(site.get("phone")),                  "Téléphone affiché (avec espaces)"),
        ("phoneHref",             s(site.get("phoneHref")),              "Téléphone SANS espaces pour les liens tel:"),
        ("address",               s(site.get("address")),                "Adresse complète"),
        ("logo",                  s(site.get("logo")),                   "Chemin du logo principal"),
        ("footerDescription",     s(site.get("footerDescription")),      "Description footer en français"),
        ("footerDescriptionEn",   s(site.get("footerDescriptionEn")),    "Description footer en anglais"),
        ("instagramUrl",          s(site.get("instagramUrl")),           "Lien Instagram BDE principal"),
        ("instagramLabel",        s(site.get("instagramLabel")),         "Label affiché pour l'Instagram BDE"),
        ("nuitsInstagramUrl",     s(site.get("nuitsInstagramUrl")),      "Lien Instagram Les Nuits Dauphine"),
        ("nuitsInstagramLabel",   s(site.get("nuitsInstagramLabel")),    "Label affiché pour l'Instagram Nuits"),
        ("tiktokUrl",             s(site.get("tiktokUrl")),              "Lien TikTok"),
        ("facebookUrl",           s(site.get("facebookUrl")),            "Lien Facebook"),
        ("defaultTicketUrl",      s(site.get("defaultTicketUrl")),       "Lien billetterie générique (Shotgun)"),
        ("defaultGooglePhotosUrl",s(site.get("defaultGooglePhotosUrl")), "Lien Google Photos générique"),
        ("copyrightText",         s(site.get("copyrightText")),          "Texte copyright du footer"),
        ("ticket_eventName",      s(ticket.get("eventName")),            "Nom de l'événement dans la modal billetterie"),
        ("ticket_eventSub",       s(ticket.get("eventSub")),             "Sous-titre modal billetterie (FR)"),
        ("ticket_eventSubEn",     s(ticket.get("eventSubEn")),           "Sous-titre modal billetterie (EN)"),
        ("ticket_url",            s(ticket.get("url")),                  "URL du lien billetterie générique"),
        ("ticket_buttonLabel",    s(ticket.get("buttonLabel")),          "Texte du bouton billetterie (FR)"),
        ("ticket_buttonLabelEn",  s(ticket.get("buttonLabelEn")),        "Texte du bouton billetterie (EN)"),
        ("ticket_description",    s(ticket.get("description")),          "Description billetterie courte (FR)"),
        ("ticket_descriptionEn",  s(ticket.get("descriptionEn")),        "Description billetterie courte (EN)"),
        ("ticket_note",           s(ticket.get("note")),                 "Note complémentaire billetterie (FR)"),
        ("ticket_noteEn",         s(ticket.get("noteEn")),               "Note complémentaire billetterie (EN)"),
    ]

    for i, (champ, valeur, desc) in enumerate(rows, 2):
        ws.cell(i, 1, champ).font = KEY_FONT
        ws.cell(i, 2, valeur)
        ws.cell(i, 3, desc).font = NOTE_FONT
        paint_row(ws, i, 3)


def sheet_navigation(wb, site):
    ws = wb.create_sheet("Navigation")
    make_header(ws, [
        ("ORDRE", 8),
        ("LABEL FR", 22),
        ("LABEL EN", 22),
        ("HREF", 32),
        ("LIEN HORS HOME (externalFromNonHome)", 36),
    ])
    for i, nav in enumerate(site.get("navigation") or [], 2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, s(nav.get("label")))
        ws.cell(i, 3, s(nav.get("labelEn")))
        ws.cell(i, 4, s(nav.get("href")))
        ws.cell(i, 5, s(nav.get("externalFromNonHome")))
        paint_row(ws, i, 5)


def sheet_events(wb, events):
    ws = wb.create_sheet("Events")
    cols = [
        ("SLUG",                                14),
        ("ORDRE",                               7),
        ("NUMÉRO",                              8),
        ("TITRE",                               22),
        ("SOUS-TITRE",                          32),
        ("CATÉGORIE",                           18),
        ("DATE (AAAA-MM-JJ)",                   16),
        ("DATE AFFICHÉE FR",                    20),
        ("DATE AFFICHÉE EN",                    20),
        ("STATUT FR",                           16),
        ("STATUT EN",                           16),
        ("COULEUR STATUT",                      14),
        ("LIEU",                                12),
        ("SALLE / VENUE",                       16),
        ("DESCRIPTION COURTE",                  32),
        ("HOME DESC FR",                        36),
        ("HOME DESC EN",                        36),
        ("HOME PÉRIODE FR",                     18),
        ("HOME PÉRIODE EN",                     18),
        ("IMAGE PRINCIPALE",                    32),
        ("IMAGES SUPPLÉMENTAIRES (séparer ;)",  42),
        ("ALT IMAGE",                           22),
        ("TICKET URL",                          32),
        ("TICKET LABEL",                        15),
        ("DOSSIER URL",                         32),
        ("DOSSIER LABEL FR",                    20),
        ("DOSSIER LABEL EN",                    20),
        ("DOSSIER DOWNLOAD (OUI/NON)",          22),
        ("GOOGLE PHOTOS URL",                   32),
        ("PAGE GALERIE",                        22),
        ("GALERIE LABEL",                       26),
        ("SUR HOME (OUI/NON)",                  16),
        ("SUR ÉVÉNEMENTS (OUI/NON)",            20),
        ("SUR PAGE NUITS (OUI/NON)",            20),
        ("FEATURED (OUI/NON)",                  15),
        ("STATUS",                              12),
        ("REVERSE (OUI/NON)",                   15),
    ]
    make_header(ws, cols)

    for i, ev in enumerate(events, 2):
        dl   = ev.get("dateLabel")     or {}
        sl   = ev.get("statusLabel")   or {}
        hd   = ev.get("homeDescriptionI18n") or {}
        hp   = ev.get("homePeriodI18n") or {}
        dlab = ev.get("dossierLabel")  or {}
        imgs = ev.get("images")        or []
        main = s(ev.get("image"))
        supp = " ; ".join(img for img in imgs if img != main)

        row = [
            s(ev.get("slug")),
            ev.get("order", i - 1),
            s(ev.get("number")),
            s(ev.get("title")),
            s(ev.get("subtitle")),
            s(ev.get("category")),
            s(ev.get("date")),
            s(dl.get("fr")),
            s(dl.get("en")),
            s(sl.get("fr")),
            s(sl.get("en")),
            s(ev.get("statusColor")),
            s(ev.get("place")),
            s(ev.get("venue")),
            s(ev.get("shortDescription")),
            s(hd.get("fr")),
            s(hd.get("en")),
            s(hp.get("fr")),
            s(hp.get("en")),
            main,
            supp,
            s(ev.get("alt")),
            s(ev.get("ticketUrl")),
            s(ev.get("ticketLabel")),
            s(ev.get("dossierUrl")),
            s(dlab.get("fr")),
            s(dlab.get("en")),
            b(ev.get("dossierDownload")),
            s(ev.get("googlePhotosUrl")),
            s(ev.get("galleryPage")),
            s(ev.get("galleryLabel")),
            b(ev.get("showOnHome")),
            b(ev.get("showOnEventsPage")),
            b(ev.get("showOnNuitsPage")),
            b(ev.get("featured")),
            s(ev.get("status")),
            b(ev.get("reverse")),
        ]
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        paint_row(ws, i, len(cols))


def sheet_event_descriptions(wb, events):
    ws = wb.create_sheet("Event_Descriptions")
    make_header(ws, [
        ("SLUG ÉVÉNEMENT", 16),
        ("ORDRE",           7),
        ("TEXTE FR",       70),
        ("TEXTE EN",       70),
    ])
    ri = 2
    for ev in events:
        slug = ev.get("slug", "")
        for k, block in enumerate(ev.get("descriptionBlocks") or [], 1):
            i18n = block.get("i18n") or {}
            ws.cell(ri, 1, slug)
            ws.cell(ri, 2, k)
            c3 = ws.cell(ri, 3, s(i18n.get("fr") or block.get("text")))
            c4 = ws.cell(ri, 4, s(i18n.get("en")))
            c3.alignment = Alignment(wrap_text=True, vertical="top")
            c4.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[ri].height = 60
            paint_row(ws, ri, 4)
            ri += 1


def sheet_event_meta(wb, events):
    ws = wb.create_sheet("Event_Meta")
    make_header(ws, [
        ("SLUG ÉVÉNEMENT", 16),
        ("ORDRE",           7),
        ("CLÉ FR",         22),
        ("CLÉ EN",         22),
        ("VALEUR FR",      32),
        ("VALEUR EN",      32),
    ])
    ri = 2
    for ev in events:
        slug = ev.get("slug", "")
        for k, meta in enumerate(ev.get("meta") or [], 1):
            ki = meta.get("keyI18n")   or {}
            vi = meta.get("valueI18n") or {}
            ws.cell(ri, 1, slug)
            ws.cell(ri, 2, k)
            ws.cell(ri, 3, s(ki.get("fr") or meta.get("key")))
            ws.cell(ri, 4, s(ki.get("en") or meta.get("key")))
            ws.cell(ri, 5, s(vi.get("fr") or meta.get("value")))
            ws.cell(ri, 6, s(vi.get("en") or meta.get("value")))
            paint_row(ws, ri, 6)
            ri += 1


def sheet_event_tags(wb, events):
    ws = wb.create_sheet("Event_Tags")
    make_header(ws, [
        ("SLUG ÉVÉNEMENT",  16),
        ("ORDRE",            7),
        ("LABEL",           22),
        ("ROUGE (OUI/NON)", 15),
        ("LABEL FR",        22),
        ("LABEL EN",        22),
    ])
    ri = 2
    for ev in events:
        slug = ev.get("slug", "")
        for k, tag in enumerate(ev.get("tags") or [], 1):
            i18n = tag.get("i18n") or {}
            ws.cell(ri, 1, slug)
            ws.cell(ri, 2, k)
            ws.cell(ri, 3, s(tag.get("label")))
            ws.cell(ri, 4, b(tag.get("red", False)))
            ws.cell(ri, 5, s(i18n.get("fr") or tag.get("label")))
            ws.cell(ri, 6, s(i18n.get("en")))
            paint_row(ws, ri, 6)
            ri += 1


def sheet_event_artists(wb, events):
    ws = wb.create_sheet("Event_Artists")
    make_header(ws, [
        ("SLUG ÉVÉNEMENT",       16),
        ("ORDRE",                 7),
        ("LABEL",                28),
        ("MIS EN AVANT (OUI/NON)",20),
    ])
    ri = 2
    for ev in events:
        slug = ev.get("slug", "")
        for k, art in enumerate(ev.get("artists") or [], 1):
            ws.cell(ri, 1, slug)
            ws.cell(ri, 2, k)
            ws.cell(ri, 3, s(art.get("label")))
            ws.cell(ri, 4, b(art.get("highlight", False)))
            paint_row(ws, ri, 4)
            ri += 1


def sheet_sponsors(wb, sponsors):
    ws = wb.create_sheet("Sponsors")
    make_header(ws, [
        ("NOM",                 22),
        ("LOGO (chemin)",       32),
        ("FALLBACK (initiales)",14),
        ("ALT",                 22),
        ("URL",                 32),
        ("CATÉGORIE",           20),
        ("TYPE FR",             24),
        ("TYPE EN",             24),
        ("LIEN LABEL",          24),
        ("ACTIF (OUI/NON)",     14),
        ("ORDRE",                7),
    ])
    for i, sp in enumerate(sponsors, 2):
        row = [
            s(sp.get("name")),
            s(sp.get("logo")),
            s(sp.get("fallback")),
            s(sp.get("alt")),
            s(sp.get("url")),
            s(sp.get("category")),
            s(sp.get("type")),
            s(sp.get("typeEn")),
            s(sp.get("linkLabel")),
            b(sp.get("active")),
            sp.get("order", i - 1),
        ]
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        paint_row(ws, i, 11)


def sheet_artists_cartes(wb, artists):
    ws = wb.create_sheet("Artists_Cartes")
    make_header(ws, [
        ("NOM",                   20),
        ("IMAGE (chemin)",        32),
        ("ALT",                   20),
        ("ANNÉE · ÉVÉNEMENT",     22),
        ("TEXTE ÉVÉNEMENT",       32),
        ("BADGE",                 16),
        ("FEATURED (OUI/NON)",    16),
        ("ACTIF (OUI/NON)",       14),
        ("ORDRE",                  7),
    ])
    for i, card in enumerate(artists.get("cards") or [], 2):
        row = [
            s(card.get("name")),
            s(card.get("image")),
            s(card.get("alt")),
            s(card.get("yearEvent")),
            s(card.get("eventText")),
            s(card.get("badge")),
            b(card.get("featured")),
            b(card.get("active")),
            card.get("order", i - 1),
        ]
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        paint_row(ws, i, 9)


def sheet_artists_bande(wb, artists):
    ws = wb.create_sheet("Artists_Bande")
    make_header(ws, [
        ("ORDRE",                7),
        ("NOM (bande défilante)",28),
    ])
    for i, name in enumerate(artists.get("textStrip") or [], 2):
        ws.cell(i, 1, i - 1)
        ws.cell(i, 2, s(name))
        paint_row(ws, i, 2)


def sheet_galeries(wb, galleries):
    ws = wb.create_sheet("Galeries")
    make_header(ws, [
        ("SLUG",               14),
        ("TITRE",              18),
        ("PAGE HTML",          24),
        ("SLUG ÉVÉNEMENT LIÉ", 18),
        ("IMAGE COUVERTURE",   32),
        ("GOOGLE PHOTOS URL",  36),
        ("MODE LIGHTBOX",      14),
        ("ACTIF (OUI/NON)",    14),
        ("ORDRE",               7),
    ])
    for i, gal in enumerate(galleries, 2):
        row = [
            s(gal.get("slug")),
            s(gal.get("title")),
            s(gal.get("page")),
            s(gal.get("eventSlug")),
            s(gal.get("coverImage")),
            s(gal.get("googlePhotosUrl")),
            s(gal.get("lightboxMode")),
            b(gal.get("active")),
            gal.get("order", i - 1),
        ]
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
        paint_row(ws, i, 9)


def sheet_galerie_images(wb, galleries):
    ws = wb.create_sheet("Galerie_Images")
    make_header(ws, [
        ("SLUG GALERIE", 14),
        ("ORDRE",         7),
        ("SRC (chemin)", 32),
        ("ALT",          26),
        ("LÉGENDE FR",   42),
        ("LÉGENDE EN",   42),
        ("TAG FR",       22),
        ("TAG EN",       22),
    ])
    ri = 2
    for gal in galleries:
        slug = gal.get("slug", "")
        for k, img in enumerate(gal.get("images") or [], 1):
            ci = img.get("captionI18n") or {}
            ti = img.get("tagI18n")     or {}
            ws.cell(ri, 1, slug)
            ws.cell(ri, 2, k)
            ws.cell(ri, 3, s(img.get("src")))
            ws.cell(ri, 4, s(img.get("alt")))
            ws.cell(ri, 5, s(ci.get("fr") or img.get("caption")))
            ws.cell(ri, 6, s(ci.get("en")))
            ws.cell(ri, 7, s(ti.get("fr") or img.get("tag")))
            ws.cell(ri, 8, s(ti.get("en")))
            paint_row(ws, ri, 8)
            ri += 1


# ─────────────────────────────────────────────────────────────────
# Point d'entrée
# ─────────────────────────────────────────────────────────────────

def main():
    print()
    print("  Lecture des fichiers data/*.js...")

    site      = load_js("site.js")
    events    = load_js("events.js")
    sponsors  = load_js("sponsors.js")
    artists   = load_js("artists.js")
    galleries = load_js("galleries.js")

    print("  Construction du fichier Excel...")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheet_legende(wb)
    sheet_config(wb, site)
    sheet_navigation(wb, site)
    sheet_events(wb, events)
    sheet_event_descriptions(wb, events)
    sheet_event_meta(wb, events)
    sheet_event_tags(wb, events)
    sheet_event_artists(wb, events)
    sheet_sponsors(wb, sponsors)
    sheet_artists_cartes(wb, artists)
    sheet_artists_bande(wb, artists)
    sheet_galeries(wb, galleries)
    sheet_galerie_images(wb, galleries)

    wb.save(EXCEL)

    sheets = [ws.title for ws in wb.worksheets]
    print(f"\n  OK  admin.xlsx cree : {EXCEL}")
    print(f"  Feuilles : {', '.join(sheets)}")
    print(f"  Evenements : {len(events)}  |  Sponsors : {len(sponsors)}  |  Galeries : {len(galleries)}")
    print()


if __name__ == "__main__":
    main()
